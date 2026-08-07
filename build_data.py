import openpyxl, json, sys, os
from collections import defaultdict
from datetime import datetime

# Usage: python3 build_data.py ["path/to/Department KPIs.xlsx"]
# Defaults to "Department KPIs.xlsx" in the same folder as this script.
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SRC = sys.argv[1] if len(sys.argv) > 1 else os.path.join(SCRIPT_DIR, "Department KPIs.xlsx")
OUT_PATH = os.path.join(SCRIPT_DIR, "data.json")

if not os.path.exists(SRC):
    print(f"Could not find workbook at: {SRC}")
    print("Pass the path explicitly: python3 build_data.py \"/path/to/Department KPIs.xlsx\"")
    sys.exit(1)

wb = openpyxl.load_workbook(SRC, data_only=True)

def iso(d):
    return d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else None

def iso_dt(d):
    return d.strftime("%Y-%m-%dT%H:%M:%S") if hasattr(d, "strftime") else None

out = {"generatedAt": datetime.now().strftime("%Y-%m-%d")}

# =====================================================================
# Planning Visits -- already monthly grain in the source sheet
# =====================================================================
ws = wb["Planning Visits"]
pv = []
for r in range(4, ws.max_row + 1):
    d = ws.cell(row=r, column=1).value
    if not d:
        continue
    pv.append({
        "date": iso(d),
        "year": d.year,
        "planningVisits": ws.cell(row=r, column=2).value,
        "clientsServiced": ws.cell(row=r, column=3).value,
        "partnersVisited": ws.cell(row=r, column=4).value,
        "conventionGroupsServiced": ws.cell(row=r, column=5).value,
        "inHouseGroupsServiced": ws.cell(row=r, column=6).value,
    })
out["planningVisits"] = pv

# =====================================================================
# Partner Referrals -- raw rows (date, staff, count) so the front end
# can slice by year / staff the same way the Power BI page does
# =====================================================================
ws = wb["Partner Referrals"]
referrals_raw = []
for r in range(4, ws.max_row + 1):
    d = ws.cell(row=r, column=1).value
    n = ws.cell(row=r, column=2).value
    staff = ws.cell(row=r, column=3).value
    if d and n is not None:
        referrals_raw.append({"date": iso(d), "year": d.year, "staff": staff, "count": n})
out["partnerReferrals"] = {"raw": referrals_raw}

# =====================================================================
# ACC Survey ("Visit Anaheim Team Experience" section)
# raw rows carry question, rating, date, feedback, and the DS&E manager
# on that response (column I) -- this is what the PBI "Avg Rating by
# Staff" chart on the Client Survey page groups by.
# =====================================================================
ws = wb["ACC Survey"]
Q_ORDER = []
seen = set()
for r in range(5, ws.max_row + 1):
    q = ws.cell(row=r, column=5).value
    if q and q not in seen:
        seen.add(q)
        Q_ORDER.append(q)
Q_NUM = {q: i + 1 for i, q in enumerate(Q_ORDER)}

survey_raw = []
for r in range(5, ws.max_row + 1):
    q = ws.cell(row=r, column=5).value
    sd = ws.cell(row=r, column=1).value
    if not q or not sd:
        continue
    # Per explicit instruction: "date"/"year" for the ACC Survey table are
    # Start Date (column A) everywhere on this dashboard (Client Survey tab's
    # Year filter/charts/YoY tables, the Q2/Q7 spotlight, and the Overview
    # tab's "VA Team Experience Rating" card all use this same field now).
    # "Recorded Date" (column D, when the response was submitted) is kept
    # separately as "recordedDate" in case it's needed for reference later,
    # but nothing on the dashboard filters by it as of this build.
    rd = ws.cell(row=r, column=4).value
    survey_raw.append({
        "question": q,
        "qNum": Q_NUM[q],
        "rating": ws.cell(row=r, column=6).value,
        "date": iso(sd),
        "year": sd.year,
        "recordedDate": iso_dt(rd),
        "feedback": (str(ws.cell(row=r, column=7).value).strip() if ws.cell(row=r, column=7).value else None),
        # Column H, added after this dashboard was first built: a manual
        # Positive/Neutral/Negative sentiment tag on each Q7 testimonial row,
        # feeding the Feedback section's sentiment badges/scale directly
        # instead of the earlier keyword-heuristic estimate.
        "sentiment": (str(ws.cell(row=r, column=8).value).strip() if ws.cell(row=r, column=8).value else None),
        # NOTE: adding the Sentiment column (H) shifted every column after it
        # one to the right -- Event Attendance is now column I (9), the DS&E
        # manager name is now column J (10, was 9), and LeadID is now column K
        # (11, was 10). Manager/LeadID below were updated to match; if more
        # columns are added/removed from this sheet in the future, re-check
        # these indices against the sheet's actual header row.
        "manager": ws.cell(row=r, column=10).value,
        "leadId": ws.cell(row=r, column=11).value,
    })
out["accSurvey"] = {"questions": Q_ORDER, "raw": survey_raw}

# Q2 & Q7 relation kept as a bonus "spotlight" (not present in the source
# Power BI report, but explicitly requested for this dashboard): Q2 is the
# numeric DS&E Manager rating, Q7 is the open-ended testimonial question.
q2_text = Q_ORDER[1]
q7_text = Q_ORDER[6]
YEARS = [2023, 2024, 2025, 2026]

def yearly_avg(question):
    buckets = defaultdict(list)
    for row in survey_raw:
        if row["question"] == question and row["rating"] is not None:
            buckets[row["year"]].append(row["rating"])
    return {str(y): (round(sum(buckets[y]) / len(buckets[y]), 2) if buckets[y] else None) for y in YEARS}

def yoy(a, b):
    if a is None or b is None or a == 0:
        return None
    return round((b - a) / a, 4)

q2_yearly = yearly_avg(q2_text)
q2_yoy = {f"{a}-{b}": yoy(q2_yearly[str(a)], q2_yearly[str(b)]) for a, b in zip(YEARS, YEARS[1:])}

testimonials_by_year = defaultdict(list)
for row in survey_raw:
    if row["question"] == q7_text and row["feedback"]:
        testimonials_by_year[row["year"]].append({"date": row["date"][:10], "feedback": row["feedback"]})

out["accSurvey"]["q2q7"] = {
    "q2Text": q2_text,
    "q7Text": q7_text,
    "q2Yearly": q2_yearly,
    "q2Yoy": q2_yoy,
    "testimonialSamples": {str(y): testimonials_by_year[y][:3] for y in YEARS},
    "totalTestimonials": sum(len(v) for v in testimonials_by_year.values()),
}

# =====================================================================
# Repeating ACC Clients Services
# =====================================================================
ws = wb["Repeating ACC Clients Services "]
repeat_raw = []
for r in range(5, ws.max_row + 1):
    mgr = ws.cell(row=r, column=6).value
    if mgr is None:
        continue
    start = ws.cell(row=r, column=8).value
    repeat_raw.append({
        "accountId": ws.cell(row=r, column=1).value,
        "accountName": ws.cell(row=r, column=2).value,
        "leadId": ws.cell(row=r, column=3).value,
        "leadName": ws.cell(row=r, column=4).value,
        "salesManager": ws.cell(row=r, column=5).value,
        "servicesManager": mgr,
        "conventionCenter": ws.cell(row=r, column=7).value,
        "startDate": iso(start),
        "year": start.year if hasattr(start, "year") else None,
        "endDate": iso(ws.cell(row=r, column=9).value),
        "attendance": ws.cell(row=r, column=10).value,
        "peakRoom": ws.cell(row=r, column=11).value,
        "repeat": ws.cell(row=r, column=12).value,
        "status": ws.cell(row=r, column=13).value,
    })
out["repeatingClients"] = {"raw": repeat_raw}

# =====================================================================
# Events (shared travel / events calendar)
# =====================================================================
ws = wb["Events"]
events_raw = []
skipped_bad_dates = 0
for r in range(4, ws.max_row + 1):
    d = ws.cell(row=r, column=4).value
    if not isinstance(d, datetime):
        continue
    if not (2020 <= d.year <= 2030):
        skipped_bad_dates += 1  # e.g. a "2206" typo in the source sheet
        continue
    events_raw.append({
        "event": ws.cell(row=r, column=1).value,
        "type": ws.cell(row=r, column=2).value,
        "department": (str(ws.cell(row=r, column=3).value).strip() if ws.cell(row=r, column=3).value else None),
        "date": iso(d),
        "year": d.year,
        "lead": ws.cell(row=r, column=5).value,
        "supported": ws.cell(row=r, column=6).value,
    })
out["events"] = {"raw": events_raw, "skippedInvalidDates": skipped_bad_dates}

# =====================================================================
# Booked Business -- grain is one row per ATTENDEE under a LEAD under an
# EVENT. "Leads Generated" on the Power BI page counts unique Lead IDs,
# not attendee rows -- matched by cross-checking against the live report.
# =====================================================================
ws = wb["Booked Business"]
booked_raw = []
for r in range(4, ws.max_row + 1):
    eid = ws.cell(row=r, column=1).value
    if eid is None:
        continue
    d = ws.cell(row=r, column=3).value
    booked_raw.append({
        "eventId": eid,
        "eventName": ws.cell(row=r, column=2).value,
        "eventStartDate": iso(d),
        "year": d.year if hasattr(d, "year") else None,
        "eventType": ws.cell(row=r, column=4).value,
        "eventStatus": ws.cell(row=r, column=5).value,
        "attendeeId": ws.cell(row=r, column=6).value,
        "contactId": ws.cell(row=r, column=7).value,
        "fullName": ws.cell(row=r, column=8).value,
        "accountId": ws.cell(row=r, column=9).value,
        "accountName": ws.cell(row=r, column=10).value,
        "leadId": ws.cell(row=r, column=11).value,
        "leadName": ws.cell(row=r, column=12).value,
        "salesManager": ws.cell(row=r, column=13).value,
        "leadStartDate": iso(ws.cell(row=r, column=14).value),
        "leadStatus": ws.cell(row=r, column=15).value,
        "leadCreatedDate": iso(ws.cell(row=r, column=16).value),
        "daysFromLeadCreatedToEvent": ws.cell(row=r, column=17).value,
    })
out["bookedBusiness"] = {"raw": booked_raw}

# =====================================================================
# Event Surveys
# =====================================================================
ws = wb["Event Surveys"]
event_surveys_raw = []
for r in range(4, ws.max_row + 1):
    d = ws.cell(row=r, column=5).value
    if not d:
        continue
    event_surveys_raw.append({
        "event": ws.cell(row=r, column=1).value,
        "eventId": ws.cell(row=r, column=2).value,
        "category": ws.cell(row=r, column=3).value,
        "surveyType": ws.cell(row=r, column=4).value,
        "date": iso(d),
        "year": d.year,
        "overall": ws.cell(row=r, column=6).value,
        "satisfaction": ws.cell(row=r, column=7).value,
        "registration": ws.cell(row=r, column=8).value,
        "recommend": ws.cell(row=r, column=12).value,
        "partnerName": ws.cell(row=r, column=14).value,
    })
out["eventSurveys"] = {"raw": event_surveys_raw}

with open(OUT_PATH, "w") as f:
    json.dump(out, f, indent=2, default=str)

print(f"DONE -> wrote {OUT_PATH}")
print("Questions:", Q_ORDER)
print("Planning visits rows:", len(pv))
print("Referral rows:", len(referrals_raw))
print("Survey rows:", len(survey_raw))
print("Repeat client rows:", len(repeat_raw))
print("Event rows:", len(events_raw), "skipped:", skipped_bad_dates)
print("Booked business rows:", len(booked_raw), "unique leads:", len({r['leadId'] for r in booked_raw if r['leadId']}))
print("Event survey rows:", len(event_surveys_raw))
